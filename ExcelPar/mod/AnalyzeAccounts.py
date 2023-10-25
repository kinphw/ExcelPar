import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment, numbers
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import Color, PatternFill, Font, Border
import string
import pandas as pd
import numpy as np
import tqdm

from ExcelPar.mod.SetGlobal import SetGlobal

class AnalyzeAccounts:

    @classmethod
    def AnalyzeAccounts(cls, gl:pd.DataFrame):

        #####################################################################
        ## 2. 계정분석

        # List of account codes to analyze
        account_codes = SetGlobal.분석계정과목 #Global
        SetGlobal.검토문장 = []

        account_code = account_codes[0]

        ### 진행률
        pbar = tqdm.tqdm(total=len(account_codes), desc="...")
        pbar.set_description("순환 START")

        for account_code in account_codes:

            pbar.set_description(account_code)        
            pbar.update(1)        
            
            df = gl #인수로 GL을 받아서 가공한다.

            #Preprocess
            df['전표번호'] = df['전표번호'].astype(str)
            df['계정과목코드'] = df['계정과목코드'].astype(str)
                
            df1 = df[df["Company code"] == account_code] #GL 중 대상 계정과목만 추출 => df1
            d = pd.DataFrame(np.zeros((12, 0))) #[12,0] DF 생성 => d
            d["회계월"] = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]  #d는 1~12 월 Series임
            df2 = df1.pivot_table(index=["회계월"], columns='연도', values='전표금액', aggfunc='sum').reset_index() #df1을 (행)회계월/(열)연도로 피벗함
            df2 = pd.merge(d, df2, on="회계월", how="left").fillna(0) #d & df2 join함. 그냥 df2랑 같은 것임 (1~12월의 12행)

            #당기 혹은 전기가 없는 에러 수정    
            if 'CY' not in df2.columns:
                df2['CY'] = 0
                    # "PY" 열 추가
            if 'PY' not in df2.columns:
                df2['PY'] = 0
                
            df2 = df2[["회계월", "CY", "PY"]]
            
            ########################################################################################
            # 상하단 Logic 변경 # RANGE : NEW LOGIC _ 231023 - 이예린SM
            ########################################################################################

            df_melted = df2.melt(id_vars=['회계월'], value_vars=['CY', 'PY'], var_name='구분', value_name='금액') #Melted Dataframe
            df_mean = df_melted[df_melted['금액'] != 0]["금액"].mean() #Melted df 중 금액이 있는 월의 평균
            df_std = df_melted[df_melted['금액'] != 0]["금액"].std() #표준편차            

            # Z-Score 임계값 설정 (예시: ±2)
            상단기준_Z = 2
            하단기준_Z = -2

            # 상단 기준금액과 하단 기준금액 계산
            상단금액기준 = df_mean + 상단기준_Z * df_std
            하단금액기준 = df_mean + 하단기준_Z * df_std

            df2["상단"] = 상단금액기준
            df2["하단"] = 하단금액기준
            
            df2["당기_분석대상"] = np.where(df2["CY"].abs() <= SetGlobal.De_minimis, "X",
                                    np.where((df2["CY"] < df2["상단"]) & (df2["CY"] > df2["하단"]), "X", "O")) ##########수정 10.20 
            df2["전기_분석대상"] = np.where(df2["PY"].abs() <= SetGlobal.De_minimis, "X",
                                    np.where((df2["PY"] < df2["상단"]) & (df2["PY"] > df2["하단"]), "X", "O")) ##########수정 10.20 
            
            df2 = df2.set_index('회계월')
            
            # 매출 증 전표가 원래 -인 것에 대한 반전 >> 코드 삭제함. CY, PY는 조정하면 안됨     
            #df2["CY"] = np.where(df2["CY"] < 0, df2["CY"]*-1, df2["CY"]) 
            #df2["PY"] = np.where(df2["PY"] < 0, df2["PY"]*-1, df2["PY"]) 

            #######################################################################################################
            # Save the analysis report to a separate sheet
            
            # DEBUG. 231019. / to _ (계정과목명에 /가 들어가면 안됨)
            rep = {
                '/':'_',
                ':':'_',
                '<':'_',
                '>':'_'
            }
            account_code_tmp = account_code #글자복사
            for before,after in rep.items():
                account_code_tmp = account_code_tmp.replace(before,after)

            report_filename = f"분석보고서_{SetGlobal.ClientNameDate}_{account_code_tmp}_{SetGlobal.Level}.xlsx"
            pbar.set_description(account_code + " 분석보고서 파일을 생성합니다...")
            df2.to_excel(report_filename, sheet_name="분석보고서", startrow=4) #CREATE
            wb = load_workbook(report_filename)
            sheet = wb["분석보고서"]
        
            min_column = wb.active.min_column
            max_column = wb.active.max_column
            min_row = wb.active.min_row
            max_row = wb.active.max_row

            #######################################################################################################
            # 라인 차트 생성
            line_chart = LineChart()

            data = Reference(sheet, min_col=min_column+1, max_col=max_column-2, min_row=min_row, max_row=max_row) # 111718 선급금'!$B$5:$E$17
            categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row) # 111718 선급금'!$A$6:$A$17

        
            line_chart.add_data(data, titles_from_data=True)
            line_chart.set_categories(categories)
            sheet.add_chart(line_chart, "B19") #location chart

            line_chart.title = '계정분석'
            line_chart.style = 2 #choose the chart style

            #######################################################################################################

            # formatting the report
            sheet['A1'] = '계정분석 보고서'
            sheet['A2'] = account_code
            sheet['A1'].font = Font('Arial', bold=True, size=18)
            sheet['A2'].font = Font('Arial', bold=True, size=12)

            sheet['A3'] = 'CY/PY는 해당월 순증감액입니다. (Not 잔액)'
            sheet['A4'] = '전기~당기 월별순증감액 μ±2σ(=Z-score ±2)가 상하단 정상Range이며, 해당월 순증감액이 Range를 벗어나는 경우 분석대상입니다.'

            pbar.set_description(account_code + " 분석보고서_원장_요약 시트를 생성합니다...")
        
            #######################################################################################################
            # Create a new sheet '분석보고서_원장_요약'
            wb.create_sheet("분석보고서_원장_요약", 1)
            sheet2 = wb['분석보고서_원장_요약']
            
            df3 = df2.reset_index()
            추출월 = df3[(df3["당기_분석대상"] == "O") | (df3["전기_분석대상"] == "O")]["회계월"].to_list()
            
            if len(추출월) > 0:
            
                result = pd.DataFrame()
                for i in 추출월:
                    result = pd.concat([result, df1[df1["회계월"] == i]])  
                result_pv = pd.pivot_table(result, index = ['회계월','거래처코드'],columns = ['연도'], values = ['전표금액'], aggfunc = 'sum').reset_index().fillna(0)


                #당기 혹은 전기가 없는 에러 수정
                if ( '전표금액', 'CY') not in result_pv.columns:
                    result_pv[( '전표금액', 'CY')] = 0
                # "PY" 열 추가
                if ( '전표금액', 'PY') not in result_pv.columns:
                    result_pv[( '전표금액', 'PY')] = 0

                result_pv = result_pv[[(  '회계월',   ''), ('거래처코드',   ''), ( '전표금액', 'CY'), ( '전표금액', 'PY')]]


                result_pv.columns = ["회계월", "거래처코드", "당기", "전기"]
                result_pv["차이금액"] = result_pv["당기"] - result_pv["전기"]
                result_pv["차이금액(절대값)"] = result_pv["차이금액"].abs()
                result_pv["Rank"] = result_pv.groupby(["회계월"])['차이금액(절대값)'].rank(method='min', ascending=False)


                result["월합산"] = result.groupby(["회계월","연도"])["전표금액"].transform("sum")
                result["설명율"] = result["전표금액"]/result["월합산"]
                result["기여도"] = result["설명율"].abs()
                result1 = result[result["기여도"] > 0.1].reset_index(drop=True)

                result1["전표금액(절대값)"] = result1["전표금액"].abs()
                result1["Rank"] = result1.groupby(["연도","회계월"])['전표금액(절대값)'].rank(method='min', ascending=False)
                result1 = result1[result1["Rank"] <= 3] 
                
            else:
                
                data = {
                '회계월': [0],
                '거래처코드': [0],
                '당기': [0],
                '전기': [0],
                '차이금액': [0],
                '차이금액(절대값)': [0],
                'Rank': [0]
            }

                result_pv = pd.DataFrame(data)
                
                data1 = {
                '전표번호': [0],
                '전기일자': [0],
                '연도': [0],
                '회계월': [0],
                '거래처코드': [0],
                '전표적요상세': [0],
                '전표금액': [0],
                '차변금액': [0],
                '대변금액': [0],
                '계정과목코드': [0],
                '계정과목명': [0],
                'Tier 1': [0],
                'Tier 2': [0],
                'Tier 3': [0],
                'Tier 4': [0],
                'Company code': [0],
                '월합산': [0],
                '설명율': [0],
                '기여도': [0],
                '전표금액(절대값)': [0],
                'Rank': [0]
            }
                
                result1 = pd.DataFrame(data1)
            
            
            #######################################################################################################
            # Save the data from 'result_pv' DataFrame to 'sheet2' with headers and border lines
            header_row2 = result_pv.columns.tolist()
            for r_idx, row in enumerate(result_pv.iterrows(), start=5):
                for c_idx, value in enumerate(row[1], start=1):
                    cell = sheet2.cell(row=r_idx, column=c_idx, value=value)
            
            for r_idx, header in enumerate(header_row2, start=1):
                cell = sheet2.cell(row=4, column=r_idx, value=header)
                cell.font = Font(bold=True)  # Make the header text bold
                cell.border = openpyxl.styles.Border(
                    bottom=openpyxl.styles.Side(style="thin"),
                    top=openpyxl.styles.Side(style="thin"),
                    left=openpyxl.styles.Side(style="thin"),
                    right=openpyxl.styles.Side(style="thin")
                )
                        
            #######################################################################################################
            # formatting the report
            sheet2['A1'] = '분석보고서_원장_요약'
            sheet2['A2'] = '이상변동이 확인된 월을 확인하기 위해 원장변동을 거래처별, 월별로 요약(pivot)합니다.'
            sheet2['A1'].font = Font('Arial', bold=True, size=18)
            sheet2['A2'].font = Font('Arial', bold=True, size=12)
            
            ##############################
            
            pbar.set_description(account_code + " 분석보고서_원장 시트를 생성합니다...")

            #######################################################################################################
            # Create a new sheet '분석보고서_원장'
            wb.create_sheet("분석보고서_원장", 2)  # Note: I changed the index to 2 to place it after 'sheet2'
            sheet3 = wb['분석보고서_원장']
            
            # Save the data from 'result' DataFrame to 'sheet3' with headers and border lines
            header_row3 = result1.columns.tolist()
            for r_idx, row in enumerate(result1.iterrows(), start=5):
                for c_idx, value in enumerate(row[1], start=1):
                    cell = sheet3.cell(row=r_idx, column=c_idx, value=value)
            
            for r_idx, header in enumerate(header_row3, start=1):
                cell = sheet3.cell(row=4, column=r_idx, value=header)
                cell.font = Font(bold=True)  # Make the header text bold
                cell.border = openpyxl.styles.Border(
                    bottom=openpyxl.styles.Side(style="thin"),
                    top=openpyxl.styles.Side(style="thin"),
                    left=openpyxl.styles.Side(style="thin"),
                    right=openpyxl.styles.Side(style="thin")
                )
                
        
            #######################################################################################################
            # Define cell format as a DifferentialStyle object
            fill_red = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')
            font = Font(bold=True, color='FFFFFF')
            dxf=DifferentialStyle(font=font ,fill=fill_red)

            # Generation of Rule object
            # Arg:rank sets the rank, arg:bottom selects the upper or lower rank, and arg:percent selects the rank or %.
            rule = Rule(type='top10', rank=10, bottom=False, percent=True, dxf=dxf)
            
            #wb['분석보고서_원장'].conditional_formatting.add("S5:S200", rule)
            wb['분석보고서_원장'].conditional_formatting.add("T5:T200", rule)
            wb['분석보고서_원장_요약'].conditional_formatting.add("F5:F200", rule)
            

            #https://www.shibutan-bloomers.com/python_libraly_openpyxl-7_en/5421/
            
            #max_row = wb['분석보고서_원장'].active.max_row
            for row in range(1, 200):
                wb['분석보고서']["B{}".format(row)].number_format = '#,##0'
                wb['분석보고서']["C{}".format(row)].number_format = '#,##0'
                wb['분석보고서']["D{}".format(row)].number_format = '#,##0'
                wb['분석보고서']["E{}".format(row)].number_format = '#,##0'
                wb['분석보고서']["F{}".format(row)].number_format = '#,##0'
                wb['분석보고서']["G{}".format(row)].number_format = '#,##0'
                
                wb['분석보고서'].column_dimensions["B"].width = 15
                wb['분석보고서'].column_dimensions["C"].width = 15
                wb['분석보고서'].column_dimensions["D"].width = 15
                wb['분석보고서'].column_dimensions["E"].width = 15
                wb['분석보고서'].column_dimensions["F"].width = 15
                wb['분석보고서'].column_dimensions["G"].width = 15
                wb['분석보고서'].column_dimensions["H"].width = 13
                wb['분석보고서'].column_dimensions["I"].width = 13
                
                
                wb['분석보고서_원장_요약']["C{}".format(row)].number_format = '#,##0'
                wb['분석보고서_원장_요약']["D{}".format(row)].number_format = '#,##0'
                wb['분석보고서_원장_요약']["E{}".format(row)].number_format = '#,##0'
                wb['분석보고서_원장_요약']["F{}".format(row)].number_format = '#,##0'
                wb['분석보고서_원장_요약']["G{}".format(row)].number_format = '#,##0'
                
                wb['분석보고서_원장_요약'].column_dimensions["C"].width = 15
                wb['분석보고서_원장_요약'].column_dimensions["D"].width = 15
                wb['분석보고서_원장_요약'].column_dimensions["E"].width = 15
                wb['분석보고서_원장_요약'].column_dimensions["F"].width = 15
                wb['분석보고서_원장_요약'].column_dimensions["G"].width = 15
                
                
                wb['분석보고서_원장']["G{}".format(row)].number_format = '#,##0'
                wb['분석보고서_원장']["H{}".format(row)].number_format = '#,##0'
                wb['분석보고서_원장']["I{}".format(row)].number_format = '#,##0'
                wb['분석보고서_원장']["Q{}".format(row)].number_format = '#,##0'
                wb['분석보고서_원장']["R{}".format(row)].number_format = '#,##0.00'
                wb['분석보고서_원장']["S{}".format(row)].number_format = '#,##0.00'
                wb['분석보고서_원장']["T{}".format(row)].number_format = '#,##0'
                
                wb['분석보고서_원장'].column_dimensions["B"].width = 18
                wb['분석보고서_원장'].column_dimensions["F"].width = 30
                wb['분석보고서_원장'].column_dimensions["G"].width = 15
                wb['분석보고서_원장'].column_dimensions["H"].width = 15
                wb['분석보고서_원장'].column_dimensions["I"].width = 15
                wb['분석보고서_원장'].column_dimensions["Q"].width = 15
                wb['분석보고서_원장'].column_dimensions["T"].width = 15
                
            
            # formatting the report
            sheet3['A1'] = '분석보고서_원장'
            sheet3['A2'] = '분석대상 계정의 원장 주요한 변동을 추출합니다(월별증감액합계대비증감액(기여도)이 10%이상이며, 연도별, 월별로 증감금액(절대값)이 상위 3위 이내).'
            sheet3['A3'] = '설명율 = 전표금액/월합산액, 기여도 = abs(설명율)' #231023 추가
            sheet3['A1'].font = Font('Arial', bold=True, size=18)
            sheet3['A2'].font = Font('Arial', bold=True, size=12)
            
            #sheet3.column_dimensions.group('J','S',hidden=True)     # B 부터 D 열 숨기기
            sheet3.column_dimensions.group('J','U',hidden=True)     # Hardcoded
            
            pbar.set_description(account_code + " 분석보고서 파일을 저장합니다.")

            #######################################################################################################
            # Save the workbook
            wb.save(report_filename)
                            
            #######################################################################################################
            #######################################################################################################
            #######################################################################################################
            #######################################################################################################
            #######################################################################################################
            
            #분석적검토##################################################################################################################
            #PHW주. 계정별로 수행
            pbar.set_description(account_code + " 분석보고서 문장을 생성합니다...")
            
            # 1.전반적인 개요
            분석보고서_문장 = f"당기 및 전기에 평균 변동금액대비 이상변동을 보인 월은 {', '.join(map(str, 추출월))}, 총 {len(추출월)}개 입니다."
            # 2.거래처별, 월별 요약
            월별_최대_거래처 = result_pv.groupby('회계월').apply(lambda x: x.nlargest(1, '차이금액(절대값)'))
            월별_최대_거래처.reset_index(drop=True, inplace=True)
            분석보고서_원장_요약_문장 = "해당월과 거래처로 집계한 내역에서 전기대비 가장 크게 증가한 당기 내용은 아래와 같습니다. \n"
            # 결과 문장 생성
            for index, row in 월별_최대_거래처.iterrows():
                월 = row['회계월']
                거래처 = row['거래처코드']
                금액 = int(row['차이금액'] / 1e6)  # 백만원 단위로 변환
                분석보고서_원장_요약_문장 += f"{월}월, 거래처코드: '{거래처}', {금액}백만원"
                if index < len(월별_최대_거래처) - 1:
                    분석보고서_원장_요약_문장 += "와 "
            분석보고서_원장_요약_문장 += "입니다."
            # 3. 원장세부내역 중 주요한 증감
            월별_최대_거래 = result1.groupby(['회계월']).apply(lambda x: x.nlargest(1, '전표금액(절대값)'))
            월별_최대_거래.reset_index(drop=True, inplace=True)
            분석보고서_원장_문장 = "원장을 스캔한 결과 가장 크게 증가(혹은 감소)한 월, 전표번호, 거래처, 금액은 아래와 같습니다.\n"
            for index, row in 월별_최대_거래.iterrows():
                연도 = row["연도"]
                월 = row['회계월']
                전표번호 = row['전표번호']
                거래처 = row['거래처코드']
                금액 = int(row['전표금액'] / 1e6)  # 백만원 단위로 변환
                전표적요 = row['전표적요상세']
                분석보고서_원장_문장 += f"{연도}, {월}월, 전표번호: '{전표번호}', 거래처코드: '{거래처}', {금액}백만원, 세부증가내역은 '{전표적요}'"
                if index < len(월별_최대_거래처) - 1:
                    분석보고서_원장_문장 += "와 \n"
            분석보고서_원장_문장 += "입니다."
            # 결과 문장 합치기
            
            if len(추출월) > 0:
            
                결과_문장 = f"{account_code}에서 증가(혹은 감소)한 세부 내역은 아래와 같습니다.\n"
                결과_문장 += '① ' + 분석보고서_문장 + '\n' + '② ' + 분석보고서_원장_요약_문장 + '\n' + '③ ' + 분석보고서_원장_문장
            
            else:
                결과_문장 = f"{account_code}에서 당기 및 전기에 평균 변동금액대비 이상변동을 보인 월은 없으므로 세부 분석은 수행하지 않습니다."
            

            SetGlobal.검토문장.append(결과_문장)

        pbar.close()
        print("계정순환 종료.")

    #####################################################################