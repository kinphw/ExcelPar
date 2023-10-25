#####################################################################
# Excel PAR
# v0.0.7 DD 231026
#####################################################################

# v0.0.4, 231019 : 세부계정 월별순증분석 상단/하단 Logic 변경.
# (기존) 순차기평균 + 순대기(절대값)평균  ==> 문제점. "가중"평균이 아니라 평균을 단순 합산하기때문에 Range가 과도하게 표시됨
# (수정후) 0이 아닌 월 절대값 평균
# v0.0.5, 231019 : FS Line 추가생성 구현
# v0.0.6, 231023 : 상하단 Logic 재변경_이예린SM (Mean ± 2Std)
#####################################################################
### 1. 전역변수 설정

def SetGlobal():
    global PM
    global De_minimis
    global diff_비율
    global ClientNameDate
    global Level

    #PM = 1000000000 # 10억을 기준으로 함
    PM = input("적용할 PM을 입력하세요 > ") or ' 2,300,000,000'
    try:
        PM = PM.replace(",","")
    except:
        pass
    PM = int(PM)
    print(f'입력하신 PM은 {PM:,}입니다.')

    #De_minimis = 200000000
    De_minimis = input("적용할 CTT를 입력하세요 > ") or ' 155,000,000 '
    try:
        De_minimis = De_minimis.replace(",","")
    except:
        pass
    De_minimis = int(De_minimis)
    print(f'입력하신 CTT는 {De_minimis:,}입니다.')

    ClientNameDate = input("파일명에 반영할 회사명/기준월을 입력하세요. 파일명에만 영향을 줍니다. (ex. 삼성전자2309)> ") or '케이티샛2309'

    diff_비율 = 0.2
    print(f'기본 차이비율 Threshold는 {diff_비율:.0%}')

    Level = 'Detail' #기본값 : Detail
#####################################################################

### 2-1. Import GL
def ImportGL():

    global gl

    import pandas as pd
    import numpy as np
    from mylib import myFileDialog as myfd

    glFile = myfd.askopenfilename() #PHW
    gl = pd.read_csv(glFile, encoding="utf-8-sig", sep="\t", low_memory=False)

#####################################################################

def PreprocessGL(Level : str = 'Detail'):

    import pandas as pd
    import numpy as np

    gl['전표번호'] = gl['전표번호'].astype(str) #전표번호 2 String
    gl['전기일자'] = pd.to_datetime(gl['전기일자'],format="%Y-%m-%d") #전기일자 2 Datetime
    gl['계정과목코드'] = gl['계정과목코드'].astype(str) #계정과목코드 2 String   
    gl['거래처코드'] = gl['거래처코드'].fillna('NAN') #DEBUG 231026

    if Level == 'Detail': #Detail일때만 수행
        print("Detail 수준 GL을 전처리 - Company code를 계정과목코드로 지정합니다.")
        gl["Company code"] = gl["계정과목코드"].apply(str) + "_" + gl["계정과목명"].apply(str) # Company Code    

    print("GL 전처리 Done")

#####################################################################

### 2-2. Import TB

def ImportTB():

    global tb

    import pandas as pd
    from mylib import myFileDialog as myfd

    tbFile = myfd.askopenfilename() #PHW
    tb = pd.read_csv(tbFile, encoding="utf-8-sig", sep="\t")

    print("DONE")

#####################################################################

def PreprocessTB():

    global 분석계정과목

    import pandas as pd
    import numpy as np

    tb['계정과목코드'] = tb['계정과목코드'].astype(str)

    #증감금액 = CY와 PY의 차이임. PY 전처리는 별도 코드에서 수행
    tb["증감금액"] = tb["CY"] - tb["PY"]
    
    #증감비율 
    # CY,PY 모두 0이면 0
    # CY가 0이 아니고 PY는 0이면 1
    # 이외에는 변동 / PY

    tb["증감비율"] = np.where((tb["CY"] == 0) & (tb["PY"] == 0), 0,
                        np.where((tb["CY"] != 0) & (tb["PY"] == 0), 1, tb["증감금액"]/tb["PY"]))
    ###########################################################################################

    # 통제위험 & 위험평가 결과에 따라 Threshold 달리 동적으로 적용하는 코드, 현재는 미사용

    # conditions = [
    #     (tb["통제활동의존"] == "NotRelyingOnControls") & (tb["위험수준"] == "LowerRisk"),
    #     (tb["통제활동의존"] == "NotRelyingOnControls") & (tb["위험수준"] == "HigherRisk"),
    #     (tb["통제활동의존"] == "NotRelyingOnControls") & (tb["위험수준"] == "SignificantRisk"),
    #     (tb["통제활동의존"] == "RelyingOnControls") & (tb["위험수준"] == "HigherRisk"),
    #     (tb["통제활동의존"] == "RelyingOnControls") & (tb["위험수준"] == "SignificantRisk")
    # ]

    # choices = [
    #     np.minimum(tb["CY"].abs() * 0.22, PM * 0.65),
    #     np.minimum(tb["CY"].abs() * 0.15, PM * 0.45),
    #     np.minimum(tb["CY"].abs() * 0.35, PM * 0.95),
    #     np.minimum(tb["CY"].abs() * 0.25, PM * 0.9),
    #     np.minimum(tb["CY"].abs() * 0.2, PM * 0.5)
    # ]

    # tb["Threshold"] = np.select(conditions, choices, default=0)
    ###########################################################################################

    #Threshold = Min[당기잔액*20%, PM*50%]
    tb["Threshold"] = np.minimum(
        tb['CY'].abs() * 0.2,
        PM * 0.5)

    # 변동액 <= CTT : X
    # 변동액 >= Threshold 또는 변동비율 > 20% : O

    conditions = [
        tb["증감금액"].abs() <= De_minimis,
        (tb["증감금액"].abs() >= tb["Threshold"]) | (tb["증감비율"].abs() >= diff_비율) #오타수정
    ]

    choices = ["X", "O"]

    tb['분석대상'] = np.select(conditions, choices, default="X")

    tb_분석대상 = tb[tb["분석대상"] == "O"]
    분석계정과목 = tb_분석대상["Company code"].unique() 
    #PHW주. Company code가 PAR 분석대상 계정과목 수준임. 기본은 실계정으로 전처리하여 수행해볼 것

    ## 임시파일 생성부
    ## 대분류 증감요인

    tb["T1_증감금액"] = tb.groupby(["T1"])["증감금액"].transform("sum")
    tb["T2_증감금액"] = tb.groupby(["T2"])["증감금액"].transform("sum")
    tb["T1_설명비율"] = np.where((tb["T2_증감금액"] == 0) & (tb["T1_증감금액"] == 0), 0,
                        np.where((tb["T2_증감금액"] != 0) & (tb["T1_증감금액"] == 0), 1, tb["T2_증감금액"]/tb["T1_증감금액"]))

    tb["T2_설명비율"] = np.where((tb["증감금액"] == 0) & (tb["T2_증감금액"] == 0), 0,
                        np.where((tb["증감금액"] != 0) & (tb["T2_증감금액"] == 0), 1, tb["증감금액"]/tb["T2_증감금액"]))


    #tb.to_excel("삭제.xlsx")

    #print("\t임시파일을 생성했습니다.")

def AnalyzeAccounts():

    #####################################################################

    ## 2. 계정분석

    global 검토문장
    global Level

    import openpyxl
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.styles import NamedStyle, Font, Alignment, numbers
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.styles import Color, PatternFill, Font, Border
    import string

    import pandas as pd
    import numpy as np

    import tqdm

    # List of account codes to analyze
    account_codes = 분석계정과목 #Global

    검토문장 = []

    account_code = account_codes[0]

    ### PHW append
    pbar = tqdm.tqdm(total=len(account_codes), desc="...")
    pbar.set_description("순환 START")

    for account_code in account_codes:

        pbar.set_description(account_code)        
        pbar.update(1)        
        
        df = gl

        #Preprocess
        df['전표번호'] = df['전표번호'].astype(str)
        df['계정과목코드'] = df['계정과목코드'].astype(str)
        
        #PHW주 : 매출 코드는 의미가 없어 보여 삭제
        #if "외상매출금" in account_code:
        if 1 == 0:
            
            pass

            # df1 = df[df["Company code"] == account_code]
            # df_매출 = df[df["Tier 3"] == "411000 매출"]

            # d = pd.DataFrame(np.zeros((12, 0)))
            # d["회계월"] = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]

            # df2 = df1.pivot_table(index=["회계월"], columns='연도', values='전표금액', aggfunc='sum').reset_index()
            # df_매출_pv = df_매출.pivot_table(index=["회계월"], columns='연도', values='전표금액', aggfunc='sum').reset_index()
            # df_매출_pv.columns = ["회계월", "CY_매출", "PY_매출"]

            # df2 = pd.merge(d, df2, on="회계월", how="left").fillna(0)
            # df2 = pd.merge(df_매출_pv, df2, on="회계월", how="left").fillna(0)

            # # CY 열이 양수인 값과 음수인 값 추출 후 절댓값 평균 계산
            # positive_mean = df2[df2["CY"] > 0]["CY"].abs().mean()
            # negative_mean = df2[df2["CY"] < 0]["CY"].abs().mean()

            # # 값이 NaN이면 대체값으로 df2["CY"].abs().mean() 사용
            # if np.isnan(positive_mean) or np.isnan(negative_mean):
            #     df2["상단"] = df2["CY"].abs().mean()
            # else:
            #     df2["상단"] = positive_mean + negative_mean

            # df2["하단"] = df2["상단"] * -1
            # df2["당기_분석대상"] = np.where(df2["CY"].abs() <= De_minimis, "X",
            #                         np.where(df2["CY"].abs() < df2["상단"].abs(), "X", "O"))
            # df2["전기_분석대상"] = np.where(df2["PY"].abs() <= De_minimis, "X",
            #                         np.where(df2["PY"].abs() < df2["상단"].abs(), "X", "O"))
            # df2 = df2.set_index('회계월')
            
            # df2["CY_매출"] = df2["CY_매출"]*-1
            # df2["PY_매출"] = df2["PY_매출"]*-1            

            # # Save the analysis report to a separate sheet
            # report_filename = f"분석보고서_{account_code}.xlsx"
            # df2.to_excel(report_filename, sheet_name="분석보고서", startrow=4)

            # wb = load_workbook(report_filename)
            # sheet = wb["분석보고서"]

            # min_column = wb.active.min_column
            # max_column = wb.active.max_column
            # min_row = wb.active.min_row
            # max_row = wb.active.max_row

            # # 라인 차트 생성
            # line_chart = LineChart()

            # data = Reference(sheet, min_col=min_column+3, max_col=max_column-2, min_row=min_row, max_row=max_row) # 111718 선급금'!$B$5:$E$17
            # categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row) # 111718 선급금'!$A$6:$A$17
            
            # line_chart.add_data(data, titles_from_data=True)
            # line_chart.set_categories(categories)
            # sheet.add_chart(line_chart, "B19") #location chart

            # line_chart.title = '계정분석'
            # line_chart.style = 2 #choose the chart style
            
            
            # # 라인 차트 생성
            # line_chart1 = LineChart()

            # data = Reference(sheet, min_col=min_column+1, max_col=max_column-4, min_row=min_row, max_row=max_row) # 111718 선급금'!$B$5:$E$17
            # categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row) # 111718 선급금'!$A$6:$A$17
            
            # line_chart1.add_data(data, titles_from_data=True)
            # line_chart1.set_categories(categories)
            # sheet.add_chart(line_chart1, "B33") #location chart

            # line_chart1.title = '계정분석_매출과 매출채권' 
            # line_chart1.style = 2 #choose the chart style
        
        else:
            
            df1 = df[df["Company code"] == account_code] #GL 중 대상 계정과목만 추출 => df1
            d = pd.DataFrame(np.zeros((12, 0))) #[12,0] DF 생성 => d
            d["회계월"] = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]  #d는 1~12 월 Series임
            df2 = df1.pivot_table(index=["회계월"], columns='연도', values='전표금액', aggfunc='sum').reset_index() #df1을 (행)회계월/(열)연도로 피벗함
            df2 = pd.merge(d, df2, on="회계월", how="left").fillna(0) #d & df2 join함. 그냥 df2랑 같은 것임 (1~12월의 12행)

            #당기 혹은 전기가 없는 에러 수정    
            if 'CY' not in df2.columns:
                df2['CY'] = 0
                    # "PY" 열 추가
            if 'PY' not in df2.columns:
                df2['PY'] = 0
                
            df2 = df2[["회계월", "CY", "PY"]]
            
            ########################################################################################
            # 상하단 Logic 변경 # RANGE : NEW LOGIC _ 231023 - 이예린SM
            ########################################################################################

            df_melted = df2.melt(id_vars=['회계월'], value_vars=['CY', 'PY'], var_name='구분', value_name='금액') #Melted Dataframe
            df_mean = df_melted[df_melted['금액'] != 0]["금액"].mean() #Melted df 중 금액이 있는 월의 평균
            df_std = df_melted[df_melted['금액'] != 0]["금액"].std() #표준편차            

            # Z-Score 임계값 설정 (예시: ±2)
            상단기준_Z = 2
            하단기준_Z = -2

            # 상단 기준금액과 하단 기준금액 계산
            상단금액기준 = df_mean + 상단기준_Z * df_std
            하단금액기준 = df_mean + 하단기준_Z * df_std

            df2["상단"] = 상단금액기준
            df2["하단"] = 하단금액기준
            
            df2["당기_분석대상"] = np.where(df2["CY"].abs() <= De_minimis, "X",
                                    np.where((df2["CY"] < df2["상단"]) & (df2["CY"] > df2["하단"]), "X", "O")) ##########수정 10.20 
            df2["전기_분석대상"] = np.where(df2["PY"].abs() <= De_minimis, "X",
                                    np.where((df2["PY"] < df2["상단"]) & (df2["PY"] > df2["하단"]), "X", "O")) ##########수정 10.20 
            
            df2 = df2.set_index('회계월')
            
            # 매출 증 전표가 원래 -인 것에 대한 반전 >> 코드 삭제함. CY, PY는 조정하면 안됨     
            #df2["CY"] = np.where(df2["CY"] < 0, df2["CY"]*-1, df2["CY"]) 
            #df2["PY"] = np.where(df2["PY"] < 0, df2["PY"]*-1, df2["PY"]) 

            # Save the analysis report to a separate sheet
            # DEBUG. 231019. / to _ (계정과목명에 /가 들어가면 안됨)
            rep = {
                '/':'_',
                ':':'_',
                '<':'_',
                '>':'_'
            }
            account_code_tmp = account_code #글자복사
            for before,after in rep.items():
                account_code_tmp = account_code_tmp.replace(before,after)

            # account_code_tmp = account_code.replace("/","_")
            # # DEBUG. 231023. : to _ (계정과목명에 :가 들어가면 안됨)
            # account_code_tmp = account_code.replace(":","_")
            # # DEBUG. 231026. : <> to _ 
            # account_code_tmp = account_code.replace("<","_")
            # account_code_tmp = account_code.replace(">","_")

            report_filename = f"분석보고서_{ClientNameDate}_{account_code_tmp}_{Level}.xlsx"

            pbar.set_description(account_code + " 분석보고서 파일을 생성합니다...")

            df2.to_excel(report_filename, sheet_name="분석보고서", startrow=4)

            wb = load_workbook(report_filename)
            sheet = wb["분석보고서"]
        
            min_column = wb.active.min_column
            max_column = wb.active.max_column
            min_row = wb.active.min_row
            max_row = wb.active.max_row

            # 라인 차트 생성
            line_chart = LineChart()

            data = Reference(sheet, min_col=min_column+1, max_col=max_column-2, min_row=min_row, max_row=max_row) # 111718 선급금'!$B$5:$E$17
            categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row) # 111718 선급금'!$A$6:$A$17

        
            line_chart.add_data(data, titles_from_data=True)
            line_chart.set_categories(categories)
            sheet.add_chart(line_chart, "B19") #location chart

            line_chart.title = '계정분석'
            line_chart.style = 2 #choose the chart style

            ### IF문 종료

        # formatting the report
        sheet['A1'] = '계정분석 보고서'
        sheet['A2'] = account_code
        sheet['A1'].font = Font('Arial', bold=True, size=18)
        sheet['A2'].font = Font('Arial', bold=True, size=12)

        sheet['A3'] = 'CY/PY는 해당월 순증감액입니다. (Not 잔액)'
        sheet['A4'] = '전기~당기 월별순증감액 μ±2σ(=Z-score ±2)가 상하단 정상Range이며, 해당월 순증감액이 Range를 벗어나는 경우 분석대상입니다.'

        pbar.set_description(account_code + " 분석보고서_원장_요약 시트를 생성합니다...")
    
        # Create a new sheet '분석보고서_원장_요약'
        wb.create_sheet("분석보고서_원장_요약", 1)
        sheet2 = wb['분석보고서_원장_요약']
        
        df3 = df2.reset_index()
        추출월 = df3[(df3["당기_분석대상"] == "O") | (df3["전기_분석대상"] == "O")]["회계월"].to_list()
        
        if len(추출월) > 0:
        
            result = pd.DataFrame()
            for i in 추출월:
                result = pd.concat([result, df1[df1["회계월"] == i]])  
            result_pv = pd.pivot_table(result, index = ['회계월','거래처코드'],columns = ['연도'], values = ['전표금액'], aggfunc = 'sum').reset_index().fillna(0)


            #당기 혹은 전기가 없는 에러 수정
            if ( '전표금액', 'CY') not in result_pv.columns:
                result_pv[( '전표금액', 'CY')] = 0
            # "PY" 열 추가
            if ( '전표금액', 'PY') not in result_pv.columns:
                result_pv[( '전표금액', 'PY')] = 0

            result_pv = result_pv[[(  '회계월',   ''), ('거래처코드',   ''), ( '전표금액', 'CY'), ( '전표금액', 'PY')]]


            result_pv.columns = ["회계월", "거래처코드", "당기", "전기"]
            result_pv["차이금액"] = result_pv["당기"] - result_pv["전기"]
            result_pv["차이금액(절대값)"] = result_pv["차이금액"].abs()
            result_pv["Rank"] = result_pv.groupby(["회계월"])['차이금액(절대값)'].rank(method='min', ascending=False)


            result["월합산"] = result.groupby(["회계월","연도"])["전표금액"].transform("sum")
            result["설명율"] = result["전표금액"]/result["월합산"]
            result["기여도"] = result["설명율"].abs()
            result1 = result[result["기여도"] > 0.1].reset_index(drop=True)

            result1["전표금액(절대값)"] = result1["전표금액"].abs()
            result1["Rank"] = result1.groupby(["연도","회계월"])['전표금액(절대값)'].rank(method='min', ascending=False)
            result1 = result1[result1["Rank"] <= 3] 
            
        else:
            
            data = {
            '회계월': [0],
            '거래처코드': [0],
            '당기': [0],
            '전기': [0],
            '차이금액': [0],
            '차이금액(절대값)': [0],
            'Rank': [0]
        }

            result_pv = pd.DataFrame(data)
            
            data1 = {
            '전표번호': [0],
            '전기일자': [0],
            '연도': [0],
            '회계월': [0],
            '거래처코드': [0],
            '전표적요상세': [0],
            '전표금액': [0],
            '차변금액': [0],
            '대변금액': [0],
            '계정과목코드': [0],
            '계정과목명': [0],
            'Tier 1': [0],
            'Tier 2': [0],
            'Tier 3': [0],
            'Tier 4': [0],
            'Company code': [0],
            '월합산': [0],
            '설명율': [0],
            '기여도': [0],
            '전표금액(절대값)': [0],
            'Rank': [0]
        }
            
            result1 = pd.DataFrame(data1)
        
        
        # Save the data from 'result_pv' DataFrame to 'sheet2' with headers and border lines
        header_row2 = result_pv.columns.tolist()
        for r_idx, row in enumerate(result_pv.iterrows(), start=5):
            for c_idx, value in enumerate(row[1], start=1):
                cell = sheet2.cell(row=r_idx, column=c_idx, value=value)
        
        for r_idx, header in enumerate(header_row2, start=1):
            cell = sheet2.cell(row=4, column=r_idx, value=header)
            cell.font = Font(bold=True)  # Make the header text bold
            cell.border = openpyxl.styles.Border(
                bottom=openpyxl.styles.Side(style="thin"),
                top=openpyxl.styles.Side(style="thin"),
                left=openpyxl.styles.Side(style="thin"),
                right=openpyxl.styles.Side(style="thin")
            )
                    
        # formatting the report
        sheet2['A1'] = '분석보고서_원장_요약'
        sheet2['A2'] = '이상변동이 확인된 월을 확인하기 위해 원장변동을 거래처별, 월별로 요약(pivot)합니다.'
        sheet2['A1'].font = Font('Arial', bold=True, size=18)
        sheet2['A2'].font = Font('Arial', bold=True, size=12)
        
        ##############################
        
        pbar.set_description(account_code + " 분석보고서_원장 시트를 생성합니다...")

        # Create a new sheet '분석보고서_원장'
        wb.create_sheet("분석보고서_원장", 2)  # Note: I changed the index to 2 to place it after 'sheet2'
        sheet3 = wb['분석보고서_원장']
        
        # Save the data from 'result' DataFrame to 'sheet3' with headers and border lines
        header_row3 = result1.columns.tolist()
        for r_idx, row in enumerate(result1.iterrows(), start=5):
            for c_idx, value in enumerate(row[1], start=1):
                cell = sheet3.cell(row=r_idx, column=c_idx, value=value)
        
        for r_idx, header in enumerate(header_row3, start=1):
            cell = sheet3.cell(row=4, column=r_idx, value=header)
            cell.font = Font(bold=True)  # Make the header text bold
            cell.border = openpyxl.styles.Border(
                bottom=openpyxl.styles.Side(style="thin"),
                top=openpyxl.styles.Side(style="thin"),
                left=openpyxl.styles.Side(style="thin"),
                right=openpyxl.styles.Side(style="thin")
            )
            
    
        # Define cell format as a DifferentialStyle object
        fill_red = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')
        font = Font(bold=True, color='FFFFFF')
        dxf=DifferentialStyle(font=font ,fill=fill_red)

        # Generation of Rule object
        # Arg:rank sets the rank, arg:bottom selects the upper or lower rank, and arg:percent selects the rank or %.
        rule = Rule(type='top10', rank=10, bottom=False, percent=True, dxf=dxf)
        
    #     wb['분석보고서_원장'].conditional_formatting.add("S5:S200", rule)
        wb['분석보고서_원장'].conditional_formatting.add("T5:T200", rule)
        wb['분석보고서_원장_요약'].conditional_formatting.add("F5:F200", rule)
        

    #     https://www.shibutan-bloomers.com/python_libraly_openpyxl-7_en/5421/
        
    #     max_row = wb['분석보고서_원장'].active.max_row
        for row in range(1, 200):
            wb['분석보고서']["B{}".format(row)].number_format = '#,##0'
            wb['분석보고서']["C{}".format(row)].number_format = '#,##0'
            wb['분석보고서']["D{}".format(row)].number_format = '#,##0'
            wb['분석보고서']["E{}".format(row)].number_format = '#,##0'
            wb['분석보고서']["F{}".format(row)].number_format = '#,##0'
            wb['분석보고서']["G{}".format(row)].number_format = '#,##0'
            
            wb['분석보고서'].column_dimensions["B"].width = 15
            wb['분석보고서'].column_dimensions["C"].width = 15
            wb['분석보고서'].column_dimensions["D"].width = 15
            wb['분석보고서'].column_dimensions["E"].width = 15
            wb['분석보고서'].column_dimensions["F"].width = 15
            wb['분석보고서'].column_dimensions["G"].width = 15
            wb['분석보고서'].column_dimensions["H"].width = 13
            wb['분석보고서'].column_dimensions["I"].width = 13
            
            
            wb['분석보고서_원장_요약']["C{}".format(row)].number_format = '#,##0'
            wb['분석보고서_원장_요약']["D{}".format(row)].number_format = '#,##0'
            wb['분석보고서_원장_요약']["E{}".format(row)].number_format = '#,##0'
            wb['분석보고서_원장_요약']["F{}".format(row)].number_format = '#,##0'
            wb['분석보고서_원장_요약']["G{}".format(row)].number_format = '#,##0'
            
            wb['분석보고서_원장_요약'].column_dimensions["C"].width = 15
            wb['분석보고서_원장_요약'].column_dimensions["D"].width = 15
            wb['분석보고서_원장_요약'].column_dimensions["E"].width = 15
            wb['분석보고서_원장_요약'].column_dimensions["F"].width = 15
            wb['분석보고서_원장_요약'].column_dimensions["G"].width = 15
            
            
            wb['분석보고서_원장']["G{}".format(row)].number_format = '#,##0'
            wb['분석보고서_원장']["H{}".format(row)].number_format = '#,##0'
            wb['분석보고서_원장']["I{}".format(row)].number_format = '#,##0'
            wb['분석보고서_원장']["Q{}".format(row)].number_format = '#,##0'
            wb['분석보고서_원장']["R{}".format(row)].number_format = '#,##0.00'
            wb['분석보고서_원장']["S{}".format(row)].number_format = '#,##0.00'
            wb['분석보고서_원장']["T{}".format(row)].number_format = '#,##0'
            
            wb['분석보고서_원장'].column_dimensions["B"].width = 18
            wb['분석보고서_원장'].column_dimensions["F"].width = 30
            wb['분석보고서_원장'].column_dimensions["G"].width = 15
            wb['분석보고서_원장'].column_dimensions["H"].width = 15
            wb['분석보고서_원장'].column_dimensions["I"].width = 15
            wb['분석보고서_원장'].column_dimensions["Q"].width = 15
            wb['분석보고서_원장'].column_dimensions["T"].width = 15
            
        
        # formatting the report
        sheet3['A1'] = '분석보고서_원장'
        sheet3['A2'] = '분석대상 계정의 원장 주요한 변동을 추출합니다(월별증감액합계대비증감액(기여도)이 10%이상이며, 연도별, 월별로 증감금액(절대값)이 상위 3위 이내).'
        sheet3['A3'] = '설명율 = 전표금액/월합산액, 기여도 = abs(설명율)' #231023 추가
        sheet3['A1'].font = Font('Arial', bold=True, size=18)
        sheet3['A2'].font = Font('Arial', bold=True, size=12)
        
        #sheet3.column_dimensions.group('J','S',hidden=True)     # B 부터 D 열 숨기기
        sheet3.column_dimensions.group('J','U',hidden=True)     # Hardcoded
        
        pbar.set_description(account_code + " 분석보고서 파일을 저장합니다.")

        # Save the workbook
        wb.save(report_filename)
                        
        #분석적검토##################################################################################################################
        #PHW주. 계정별로 수행하는 것임
        pbar.set_description(account_code + " 분석보고서 문장을 생성합니다...")
        
        # 1.전반적인 개요
        분석보고서_문장 = f"당기 및 전기에 평균 변동금액대비 이상변동을 보인 월은 {', '.join(map(str, 추출월))}, 총 {len(추출월)}개 입니다."
        # 2.거래처별, 월별 요약
        월별_최대_거래처 = result_pv.groupby('회계월').apply(lambda x: x.nlargest(1, '차이금액(절대값)'))
        월별_최대_거래처.reset_index(drop=True, inplace=True)
        분석보고서_원장_요약_문장 = "해당월과 거래처로 집계한 내역에서 전기대비 가장 크게 증가한 당기 내용은 아래와 같습니다. \n"
        # 결과 문장 생성
        for index, row in 월별_최대_거래처.iterrows():
            월 = row['회계월']
            거래처 = row['거래처코드']
            금액 = int(row['차이금액'] / 1e6)  # 백만원 단위로 변환
            분석보고서_원장_요약_문장 += f"{월}월, 거래처코드: '{거래처}', {금액}백만원"
            if index < len(월별_최대_거래처) - 1:
                분석보고서_원장_요약_문장 += "와 "
        분석보고서_원장_요약_문장 += "입니다."
        # 3. 원장세부내역 중 주요한 증감
        월별_최대_거래 = result1.groupby(['회계월']).apply(lambda x: x.nlargest(1, '전표금액(절대값)'))
        월별_최대_거래.reset_index(drop=True, inplace=True)
        분석보고서_원장_문장 = "원장을 스캔한 결과 가장 크게 증가(혹은 감소)한 월, 전표번호, 거래처, 금액은 아래와 같습니다.\n"
        for index, row in 월별_최대_거래.iterrows():
            연도 = row["연도"]
            월 = row['회계월']
            전표번호 = row['전표번호']
            거래처 = row['거래처코드']
            금액 = int(row['전표금액'] / 1e6)  # 백만원 단위로 변환
            전표적요 = row['전표적요상세']
            분석보고서_원장_문장 += f"{연도}, {월}월, 전표번호: '{전표번호}', 거래처코드: '{거래처}', {금액}백만원, 세부증가내역은 '{전표적요}'"
            if index < len(월별_최대_거래처) - 1:
                분석보고서_원장_문장 += "와 \n"
        분석보고서_원장_문장 += "입니다."
        # 결과 문장 합치기
        
        if len(추출월) > 0:
        
            결과_문장 = f"{account_code}에서 증가(혹은 감소)한 세부 내역은 아래와 같습니다.\n"
            결과_문장 += '① ' + 분석보고서_문장 + '\n' + '② ' + 분석보고서_원장_요약_문장 + '\n' + '③ ' + 분석보고서_원장_문장
        
        else:
            결과_문장 = f"{account_code}에서 당기 및 전기에 평균 변동금액대비 이상변동을 보인 월은 없으므로 세부 분석은 수행하지 않습니다."
        

        검토문장.append(결과_문장)

    pbar.close()
    print("계정순환 종료.")

#####################################################################

### 월별증감/누적월별증감을 계산하여 파일로 추출하는 부

def SummarizeMonthlyVarAmount():

    global tb_월별
    global Level

    import pandas as pd
    import numpy as np
    import openpyxl

    #############################################################

    ### 1. 월별 순변동액 (not 잔액)

    #a. 당기GL을 추출한다.
    gl_당기 = gl[gl["연도"] == "CY"] 

    #b. 당기GL을 행은 계정과목 / 열은 회계월로 피벗(SUM)한다. => TB와 JOIN
    gl_월별 = pd.pivot_table(gl_당기,values=('전표금액'),index=['계정과목코드','계정과목명'],columns=['회계월'],aggfunc=np.sum)

    # 후처리
    gl_월별 = gl_월별.fillna(0)

    #a. 월별증감액    
    Filename = f'분석보고서_계정별월별증감액_{ClientNameDate}_{Level}.xlsx'
    gl_월별.to_excel(Filename)

    # 안내문 추가
    
    wb = openpyxl.load_workbook(Filename)
    ws = wb.active
    ws.insert_rows(1)
    ws['A1'] = "#계정과목별 월별 순변동금액 - G/L로부터 추출한 금액"    
    wb.save(Filename)
    wb.close()

    print("월별 순변동액 보고서를 생성하였습니다.")

    #############################################################

    ### 2. 월말 잔액 (누적변동액)

    #a. GL은 앞에서 추출한 gl_월별 활용

    #b. TB는 그냥 가져온다. (월별이 아님)
    tb_월별 = tb[['T1', 'T2', 'T3', 'T4','Company code','통제활동의존', '위험수준', '증감금액', '증감비율', 'Threshold', '분석대상','계정과목코드', '계정과목명', 'CY', 'PY', 'PY1']]

    ### b-1. 231019 보완 => PL의 PY는 0으로 변경한다.
    tb_월별['PY'] = np.where(tb['BSPL'] == 'BS', tb['PY'], 0)    

    #c. 당기 GL(피벗) & TB를, 계정과목 KEY로 OUTER JOIN한다.
    avg_bal = pd.merge(gl_월별, tb_월별, on=['계정과목코드', '계정과목명'], how='outer')

    #d. 누적합(cumsum)을 위해 열 순서를 정렬한다.

    # 보완 => 동적으로 구현
    columns = ['T1', 'T2', 'T3', 'T4','Company code','통제활동의존', '위험수준', '증감금액', '증감비율', 'Threshold', '분석대상', '계정과목코드','계정과목명','PY']
    start_column = len(columns) - 1
    
    TmpList = gl_당기['회계월'].drop_duplicates().to_list() #당기 월수를 추출하고,
    TmpList.sort() #오름차순 정렬한다. (누적합을 위해)
    columns = columns + TmpList #합치고,
    columns.append('CY') #마지막에 당기말 포함

    # 누적합을 위한 임시 DF 생성
    avg_bal1 = avg_bal[columns]        
    end_column = len(avg_bal1.columns) -1 #Range는 마지막은 포함하지 않으므로, 그냥 CY까지 선택하면 됨

    # cumsum 수행할 열 선택
    selected_columns = avg_bal1.columns[start_column:end_column]    

    # 선택한 열에 대해서만 cumsum 수행
    avg_bal1[selected_columns] = avg_bal1[selected_columns].fillna(0).cumsum(axis=1)

    ################ 저장 ######################

    #a. 월별잔액    
    Filename = f'분석보고서_계정별월별잔액(기초+월별누적증감)_{ClientNameDate}_{Level}.xlsx'
    avg_bal1.to_excel(Filename)

    # 안내문 추가
    
    wb = openpyxl.load_workbook(Filename)
    ws = wb.active
    ws.insert_rows(1)
    ws['A1'] = "#계정과목별 각 월말 잔액(기초잔액 + 각 월 누적 증감액) - G/L로부터 추출한 금액"
    ws.column_dimensions.group('B','L',hidden=True)     
    wb.save(Filename)
    wb.close()

    print("월별 말잔 보고서를 생성하였습니다.")    
    #####################################################################

def CreateLeadReport():

    #### 총괄 분석적검토 자동화 파일을 생성합니다.

    global Level
    #tb_월별 : 전역변수 Read

    import pandas as pd
    import openpyxl
    from openpyxl.styles import NamedStyle, Font, Alignment, numbers

    분석적검토 = tb_월별[['T1', 'T2', 'T3', 'T4','Company code','통제활동의존', '위험수준','계정과목코드', '계정과목명', 'CY', 'PY', 'PY1','증감금액', '증감비율', 'Threshold', '분석대상']]
    분석적검토['PY'] = tb['PY'] #누적합을 위해 구현한 분반기코드 원복 => PY는 TB PY로
    분석적검토1 = 분석적검토[분석적검토["분석대상"] == "O"]
    # 분석적검토

    분석적검토1["주요증감"] = ""
    분석적검토1["Refer to 주요증감_파일"] = ""

    분석적검토1["주요증감"]= 검토문장 #Global
    #231025 DEBUG : 만약, TB COA에 중복이 있는 경우 여기서 걸림. 따라서 TB COA에 중복이 있으면 안됨

    pd.DataFrame(검토문장).to_excel("b.xlsx")

    #print(분석적검토1)
    #print(검토문장)

    분석적검토1["Refer to 주요증감_파일"] = "분석보고서_" + 분석계정과목 + ".xlsx"

    # 분석적검토와 분석적검토1을 합치기
    분석적검토 = pd.concat([분석적검토, 분석적검토1], ignore_index=True)

    # 분석적검토1의 "주요증감" 열을 "검토문장" 값으로 채우기
    분석적검토["주요증감"].fillna(분석적검토1["주요증감"], inplace=True)
    분석적검토["Refer to 주요증감_파일"].fillna(분석적검토1["Refer to 주요증감_파일"], inplace=True)

    분석적검토 = 분석적검토.drop_duplicates()

    print("총괄파일을 생성합니다.")
    global LeadFileName
    LeadFileName = f"총괄_분석적검토_자동화_{ClientNameDate}_{Level}.xlsx"
    분석적검토.to_excel(LeadFileName)

    # 엑셀 파일 열기
    workbook = openpyxl.load_workbook(LeadFileName)
    worksheet = workbook.active  # 또는 원하는 시트를 선택합니다.

    # pandas 데이터프레임의 열마다 컴마 스타일 적용
    for col_num, column in enumerate(분석적검토.columns, start=1):
        for cell in worksheet.iter_rows(min_row=2, min_col=col_num, max_col=col_num):
            for cell in cell:
                cell.number_format = '#,##0'
                
    for row in range(1, 1500):
            worksheet["O{}".format(row)].number_format = '#,##0.00'
            

    for column_letter in range(ord('B'), ord('P') + 1):
        column_letter = chr(column_letter)
        worksheet.column_dimensions[column_letter].width = 15

    worksheet.column_dimensions['R'].width = 50

    # 행 삽입하기
    worksheet.insert_rows(1,4)

    # formatting the report
    worksheet['A1'] = 'Preliminary Analytical Procedures'
    worksheet['A2'] = "(유의사항: 분석적 검토를 지원하기 위해 자동화된 방법으로 원장 내용을 추출하여 수행한 결과이므로 참고용으로 활용하시기 바랍니다.)"
    worksheet['A1'].font = Font('Arial', bold=True, size=18)
    worksheet['A2'].font = Font('Arial', bold=True, size=12)

    global PM
    global De_minimis
    global diff_비율    

    # 회사명 / M / PM / CTT
    worksheet.title ='총괄'

    worksheet['I3'] = '회사명'
    worksheet['I4'] = ClientNameDate
    worksheet['J3'] = '계정과목수준'
    worksheet['J4'] = Level
    worksheet['K3'] = 'PM'
    worksheet['K4'] = PM
    worksheet['L3'] = 'CTT'
    worksheet['L4'] = De_minimis
      
    worksheet['Q2'] = '분반기검토시 PY금액은 BS의 경우 전기말, PL의 경우 전년동기말입니다. (PY1은 일괄 전전기말)'
    worksheet['Q3'] = 'Threshold = Min[0.2CY, 0.5PM]' #231025 typo
    worksheet['Q4'] = '분석대상 = 증감금액 >= Threshold | 증감비율 >= 20%'
    
    from openpyxl.styles import Alignment, PatternFill

    for row in worksheet['I3:L4']:
        for cell in row:
            cell.fill = PatternFill(start_color='666699', end_color='666699',fill_type='solid')
            cell.alignment = Alignment(horizontal="center")

    for row in worksheet['I3:L3']:
        for cell in row:
            cell.fill = PatternFill(start_color='666699', end_color='666699',fill_type='solid')
            cell.alignment = Alignment(horizontal="center")

    for row in worksheet['K4:L4']:
        for cell in row:
            cell.number_format = '#,###'    

    # 조건부 서식을 적용할 열 범위 지정 (예: F5부터 F200까지)
    cell_range = worksheet['Q5:Q1000']

    # 빨간색 배경에 하얀색 폰트 스타일 설정
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 빨간색 배경색
    white_font = Font(color="FFFFFF")  # 하얀색 폰트

    # 조건부 서식 설정
    for row in cell_range:
        for cell in row:
            if cell.value == "O":
                cell.fill = red_fill  # 빨간색 배경색 적용
                cell.font = white_font  # 하얀색 폰트 적용
                
    worksheet.column_dimensions['S'].width = 40

    worksheet.freeze_panes = 'B6'  # B6 셀을 기준으로 위쪽 행과 왼쪽 열을 틀고정
    worksheet.column_dimensions.group('B','H',hidden=True)     # B 부터 D 열 숨기기


    # B5부터 S5까지의 범위 선택
    cell_range = worksheet['B5:S5']

    # 검은색 배경에 흰색 폰트색 설정
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # 검은색 배경색
    white_font = Font(color="FFFFFF", bold=True)  # 흰색 볼드 폰트

    # 범위에 스타일 적용
    for row in cell_range:
        for cell in row:
            cell.fill = black_fill  # 검은색 배경색 적용
            cell.font = white_font  # 흰색 폰트 적용
            
    workbook.save(LeadFileName)
    workbook.close()

########################################################

# 폴더 생성 및 파일 이동 함수 선언 및 정의
def create_and_move_folder(file_extension, folder_name, keyword): #file_extension : 확장자 / folder_name : 폴더명 / keyword : 파일명에 포함할 키워드

    import os
    import shutil
    current_directory = os.getcwd()

    #폴더가 없으면 폴더를 만듬
    if not os.path.exists(folder_name):
        os.mkdir(folder_name)

    for filename in os.listdir(current_directory):
        if filename.endswith(file_extension) and keyword in filename:
            source_path = os.path.join(current_directory, filename)
            destination_path = os.path.join(folder_name, filename)
            shutil.move(source_path, destination_path)

#####################################################################

def Postprocess(Level : str = 'Detail'): #Level은 Detail이 기본값임
    import os
    from openpyxl import load_workbook

    # Excel 파일을 '분석보고서_Reporting' 폴더로 이동 => 마지막에 수행
    #create_and_move_folder('.xlsx', '분석보고서_Reporting', '분석보고서_분석적검토_자동화')

    # 월별 변동액/잔액 파일을 '분석보고서_Monthly' 폴더로 이동
    create_and_move_folder('.xlsx', '분석보고서_Monthly', '분석보고서_계정별월별')

    # Excel 파일을 '분석보고서_Reporting_세부' 폴더로 이동
    create_and_move_folder('.xlsx', '분석보고서_Reporting_세부', '분석보고서')

    # Excel 파일 로드
    FileName = LeadFileName
    wb = load_workbook(FileName)
    ws = wb.active  # 첫 번째 워크시트

    # 분석보고서_Reporting_세부 폴더의 파일 리스트 가져오기 ## 세부 폴더를 추가해서 올려야함 >>>>> PHW    
    reporting_folder_path = "분석보고서_Reporting_세부"
    file_names = os.listdir(reporting_folder_path)

    # S열 6행부터 200행까지 검사
    for row in range(5, 1500):
        cell_value = ws[f"S{row}"].value  # S열의 값을 가져옴
        if cell_value in file_names:  # 값이 폴더의 파일 이름과 동일한 경우
            ws[f"S{row}"].hyperlink = os.path.join(reporting_folder_path, cell_value)  # 링크 추가

    wb.save(FileName) 
    wb.close()

    # Excel 파일을 '분석보고서_Reporting' 폴더로 이동 => 마지막에 수행
    create_and_move_folder('.xlsx', '분석보고서_Lead', '총괄_분석적검토_자동화')

    Folder1 = "분석보고서_Monthly"
    Folder2 = "분석보고서_Reporting_세부"
    Folder3 = "분석보고서_Lead"

    import os
    try:
        os.rename(Folder1,Folder1 + "_" + Level)    
    except:
        print("폴더명 변경 오류. 이미 폴더가 있는 것 같습니다.")
    try:
        os.rename(Folder2,Folder2 + "_" + Level)    
    except:
        print("폴더명 변경 오류. 이미 폴더가 있는 것 같습니다.")
    try:
        os.rename(Folder3,Folder3 + "_" + Level)
    except:
        print("폴더명 변경 오류. 이미 폴더가 있는 것 같습니다.")

#####################################################################

def ChangeTBGL():
        
    global tb
    import numpy as np
    
    #1. GL [Company Code]를 FS Line으로 변경한다.
    global gl
    gl['Company code'] = gl['FSCode'].fillna(0).astype(float).astype(int).astype(str) + "_" + gl['FSName'].astype(str) #DEBUG 231025

    gl['계정과목코드'] = gl['FSCode'].fillna(0).astype(float).astype(int).astype(str) #DEBUG 231025
    gl['계정과목명'] = gl['FSName']

    print("TB와 GL의 KEY가 FS Line으로 변경되었습니다.")
    
    #2. TB를 FS Line 기준으로 재합산한다.    
    tbFS = tb.groupby(['FSCode', 'FSName'])[['당기말','전년동기말','전전기말','전기말']].sum()    
    dfTmp = tb[['FSCode','FSName','BSPL']].drop_duplicates() #계정과목 매핑 테이블을 스스로 생성    
    tbFS = tbFS.merge(dfTmp,how='left', on='FSCode') #BSPL까지 붙임    

    tbDetail = tb #원래 tb

    tbFS["CY"] = tbFS["당기말"]
    tbFS["PY"] = np.where(tbFS["BSPL"] == "BS", tbFS["전기말"], tbFS["전년동기말"]) #조건식으로 브로드캐스팅
    tbFS["PY1"] = tbFS["전전기말"]

    tbFS['계정과목코드'] = tbFS['FSCode'].fillna(0).astype(float).astype(int).astype(str)
    tbFS['계정과목명'] = tbFS['FSName']
    tbFS["Company code"] = tbFS["계정과목코드"].apply(str) + "_" + tbFS["계정과목명"].apply(str) # Company Code

    tbFS['T1'] = 'T1'
    tbFS['T2'] = 'T2'
    tbFS['T3'] = 'T3'
    tbFS['T4'] = 'T4'
    tbFS['통제활동의존'] = 'CR'
    tbFS['위험수준'] = 'RL'

    tb = tbFS
    #여기까지 하면 준비완료    

#####################################################################

# MAIN
if(__name__=="__main__"):

    #HANDLER
    print("###EXCEL PAR BEGIN:")

    print("\n###Phase1 : Detail")
    
    print("\n#1. 기초정보를 설정합니다.")
    SetGlobal()

    print("\n#2-1. GL을 Import합니다.")
    ImportGL()

    print("\n#2-2. GL을 전처리합니다.")
    PreprocessGL('Detail') #Level

    print("\n#3-1. TB를 Import합니다.")
    ImportTB()

    print("\n#3-2. TB를 전처리합니다.")
    PreprocessTB()    

    print("\n#4. 계정별 분석을 실시하고 계정별 분석보고서를 생성합니다.")
    AnalyzeAccounts()

    print("\n#5. 계정별 월별/누적월별증감액 보고서를 생성합니다.")
    SummarizeMonthlyVarAmount()

    print("\n#6. 총괄 분석보고서를 생성합니다.")    
    CreateLeadReport()

    print("\n#7. 후처리 후 파일을 정리합니다.")    
    Postprocess("Detail") #매개변수 구현



    print("\n\n###Phase2 : FS Line")

    Level = "FSLine" #Set g_var

    print("\n#0. TB와 GL의 계정과목을 FS Line으로 대체합니다.")    
    ChangeTBGL()

    print("\n#3-2. TB를 전처리합니다.")
    PreprocessTB()

    print("\n#4. 계정별 분석을 실시하고 계정별 분석보고서를 생성합니다.")
    AnalyzeAccounts()

    print("\n#5. 계정별 월별/누적월별증감액 보고서를 생성합니다.")
    SummarizeMonthlyVarAmount()

    print("\n#6. 총괄 분석보고서를 생성합니다.")    
    CreateLeadReport()

    print("\n#7. 후처리 후 파일을 정리합니다.")    
    Postprocess("FSLine") #매개변수 구현

    #TB를 FSLine으로 합산가공하는 Code 구현
    #GL의 Company Code를 변경하는 Code 구현
    # 4,5,6,7 재실시

    print("###EXCEL PAR END:...")        